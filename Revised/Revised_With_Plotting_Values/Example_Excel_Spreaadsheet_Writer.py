import xlsxwriter
from openpyxl import load_workbook
import os

# Create a workbook and worksheet
workbook = xlsxwriter.Workbook('example.xlsx')
worksheet = workbook.add_worksheet()


# Initialize data dictionary
data = {'AAPL': [], 'NVDA': []}

# Write data to the Excel file
row = 0
col = 0
for stock, values in data.items():
    worksheet.write(row, col, stock)  # Write stock name
    for i, record in enumerate(values):
        worksheet.write_row(row + i + 1, col, record)  # Write open, current, change
    col += 4  # Move to the next set of columns for the next stock

# Close the workbook
workbook.close()

# Set permissions for the file
path = 'example.xlsx'
os.chmod(path, 0o777)

# Read and update the Excel file
workbook = load_workbook('example.xlsx')
worksheet = workbook.active

# Read existing data into a dictionary
existing_data = {}
for col in worksheet.iter_cols(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    stock_name = col[0].value
    stock_values = [list(cell.value for cell in row) for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col[0].column, max_col=col[0].column + 2)]
    existing_data[stock_name] = stock_values

# Append new data to the existing data
new_data = {
    'AAPL': [[150, 155, 5]],  # Example new data
    'NVDA': [[450, 460, 10]]
}

for stock, values in new_data.items():
    if stock in existing_data:
        existing_data[stock].extend(values)
    else:
        existing_data[stock] = values

# Write updated data back to the Excel file
workbook = xlsxwriter.Workbook('example.xlsx')
worksheet = workbook.add_worksheet()

row = 0
col = 0
for stock, values in existing_data.items():
    worksheet.write(row, col, stock)  # Write stock name
    for i, record in enumerate(values):
        worksheet.write_row(row + i + 1, col, record)  # Write open, current, change
    col += 4  # Move to the next set of columns for the next stock

workbook.close()

print(data)