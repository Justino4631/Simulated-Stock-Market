import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

file = 'stock_data.xlsx'
workbook = load_workbook(file)
worksheet = workbook.active
workbook.close()
data = {}
for column in worksheet.iter_cols(min_row=1, values_only=True):
    data[column[0]] = [cell for cell in column[1:] if cell is not None]
for key, value in data.items():
    value = value[0]
    value = value.split(";")
    only_closes = []
    for i in range(len(value) - 1):
        only_closes.append(float(value[i].split(", ")[1]))
    data[key] = only_closes
def plot_stock_data(stock_name, percent_change):
    if stock_name in data:
        bruh = float(percent_change / 100)
        x_values = list(range(len(data[stock_name])))
        plt.plot(x_values, data[stock_name], marker='o', linestyle='--', color='indigo', linewidth=2, markersize=6)  # Prettier line and markers
        plt.title(f"Stock Data for {stock_name}", fontsize=14, fontweight='bold')
        plt.xlabel("Time", fontsize=12)
        plt.ylabel("Price", fontsize=12)
        plt.xlim(left=0, right=len(data[stock_name]) - 1)
        plt.ylim(bottom=(min((data[stock_name])) * (1-bruh)), top=(max((data[stock_name]))) * (1+bruh))
        plt.axhline(y=min(data[stock_name]) * (1-bruh), color='red', linestyle='--', label=f'Min -{percent_change}%')
        plt.xticks(x_values)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.show()
    else:
        print(f"No data available for {stock_name}")