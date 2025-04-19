from openpyxl import Workbook, load_workbook
import time

file = 'stock_data.xlsx'
workbook = load_workbook(file)
worksheet = workbook.active
important_stuff = {}

for col in worksheet.iter_cols(values_only=True):
    stock_name = col[0]
    if stock_name is None:
        continue

    most_recent_value = None

    for row in col[1:]:
        if row is not None:
            try:
                items = row.split(";")
                for item in items:
                    open_price, current_price, change = item.split(", ")
                    most_recent_value = {
                        "Open": open_price.strip(),
                        "Current": current_price.strip(),
                        "Change": change.strip() + '%'
                    }
            except ValueError:
                pass

    if most_recent_value:
        important_stuff[stock_name] = most_recent_value

workbook.close()

def read_info(stock_name):
    if stock_name in list(important_stuff.keys()):
        return important_stuff[stock_name]
    else:
        return None

def read_all_info():
    for count in range(0, len(important_stuff.keys())):
        stock_name = list(important_stuff.keys())[count]
        print(f"Stock Name: {stock_name}")
        print(f"Open Price: {important_stuff[stock_name]['Open']}")
        print(f"Current Price: {important_stuff[stock_name]['Current']}")
        print(f"Change: {important_stuff[stock_name]['Change']}")
        print("-" * 20)
        time.sleep(1.5)

def get_money():
    file = open("Available_Money.txt", "r")
    money = float(file.read())
    file.close()
    return money

def buy_stock(stock_name, amount):
    stock_file = open("Stocks.txt", "r")
    stocks = stock_file.read()
    stock_file.close()
    stuff = stocks.split(",")
    dictionary = {}
    available_money = get_money()
    for item in stuff:
        uh = item.split(":")
        if(uh[0] == ""):
            continue
        if(uh[0] == stock_name):
            dictionary[uh[0]] = int(uh[1]) + amount
            available_money -= round(float(important_stuff[stock_name]['Current']) * amount, 2)
            write_to_file(available_money)
        else:
            dictionary[uh[0]] = int(uh[1])
    with open("Stocks.txt", "w") as stock_file:
        for key in dictionary.keys():
            stock_file.write(f"{key}:{dictionary[key]},")    

def sell_stock(stock_name, amount):
    with open("Stocks.txt", "r") as stock_file:
        stocks = stock_file.read()
    stock_entries = stocks.split(",")
    stock_dict = {}
    available_money = get_money()
    for item in stock_entries:
        if item.strip() == "":
            continue
        key = item.split(":")[0]
        value = item.split(":")[1]
        if(key == stock_name):
            stock_dict[key] = int(value) - amount
            available_money += round(float(important_stuff[stock_name]['Current']) * amount, 2)
            write_to_file(available_money)
        else:
            stock_dict[key] = int(value)
    with open("Stocks.txt", "w") as stock_file:
        for key in stock_dict.keys():
            stock_file.write(f"{key}:{stock_dict[key]},")

def get_portfolio():
    stock_file = open("Stocks.txt", "r")
    stocks = stock_file.read()
    stock_file.close()
    stuff = stocks.split(",")
    dictionary = {}
    for item in stuff:
        uh = item.split(":")
        if(uh[0] == ""):
            continue
        dictionary[uh[0]] = int(uh[1])
    print(dictionary)

def write_to_file(moolah):
    with open("Available_Money.txt", "w") as file:
        file.write(str(moolah))

def get_total_portfolio_value():
    stock_file = open("Stocks.txt", "r")
    stocks = stock_file.read()
    stock_file.close()
    stuff = stocks.split(",")
    dictionary = {}
    for item in stuff:
        uh = item.split(":")
        if(uh[0] == ""):
            continue
        dictionary[uh[0]] = int(uh[1])
    total_value = 0
    for key in dictionary.keys():
        total_value += round(float(important_stuff[key]['Current']) * dictionary[key], 2)
    total_value += get_money()
    return total_value

def main():
    while True:
        try:
            choice = int(input("\nEnter 1 to read specific stock info, 2 to read all info, 3 to buy a stock, 4 to sell a stock, 5 to get available money, 6 to get portfolio, 7 to get total value, or 0 to quit: \n"))
        except ValueError:
            print("Invalid input. Please enter a number.")
            continue

        if choice == 1:
            stock_name = input("Enter the stock name: \n\n").upper()
            info = read_info(stock_name)
            if info:
                print(f"Stock Name: {stock_name}")
                print(f"Open Price: {info['Open']}")
                print(f"Current Price: {info['Current']}")
                print(f"Change: {info['Change']}")
            else:
                print("Stock not found.")
            continue
        elif choice == 2:
            print(read_all_info())
            continue
        elif choice == 3:
            stock_name = input("Enter the stock name you want to buy: \n").upper()
            amount = int(input("Enter the amount you want to buy: \n"))
            buy_stock(stock_name, amount)
            continue
        elif choice == 4:
            stock_name = input("Enter the stock name you want to sell: \n").upper()
            amount = int(input("Enter the amount you want to sell: \n"))
            sell_stock(stock_name, amount)
            continue
        elif choice == 5:
            print(f"Available Money: {get_money()}")
            continue
        elif choice == 6:
            print("Your portfolio:\n")
            get_portfolio()
            continue
        elif choice == 7:
            print(f"Total Portfolio Value: {get_total_portfolio_value()}")
            continue
        elif choice == 0:
            break
        else:
            print("Invalid choice. Please enter 1, 2, or 0.")

if __name__ == "__main__":
    while(True):
        main()
        time.sleep(3600)