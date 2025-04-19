import yfinance
import time
from openpyxl import Workbook, load_workbook

stock_list = ['AAPL', 'NVDA', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'META', 'NFLX', 'AMD', 'INTC', 'CSCO', 'IBM', 'ORCL', 'ADBE', 'SHOP']

excel_file = 'stock_data.xlsx'

try:
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    worksheet = workbook.active
    for col, stock in enumerate(stock_list, start=1):
        worksheet.cell(row=1, column=col, value=stock)
    workbook.save(excel_file)

def append_to_excel(data):
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.value = None

    for col, stock in enumerate(stock_list, start=1):
        string_to_write = ''
        latest_record = data[stock]
        for record in latest_record:
            string_to_write += f"{record[0]}, {record[1]}, {record[2]}; "
        worksheet.cell(row=2, column=col, value=string_to_write)

    workbook.save(excel_file)
    workbook.close()

data = {stock: [] for stock in stock_list}

while True:
    for stock in stock_list:
        stuff = yfinance.Ticker(stock).history(period="1d", interval="60m")
        individual_data = {}
        individual_data["Open"] = round(float(stuff['Open'].iloc[0]), 2)
        individual_data["Current"] = round(float(stuff['Close'].iloc[-1]), 2)
        individual_data["Change"] = str(round(100 * (individual_data["Current"] - individual_data["Open"]) / individual_data["Open"], 2))

        data[stock].append([individual_data["Open"], individual_data["Current"], individual_data["Change"]])

    append_to_excel(data)

    print(15 * "\n")
    print(data)
    time.sleep(3600)