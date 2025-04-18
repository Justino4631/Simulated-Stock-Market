from openpyxl import Workbook, load_workbook
import time

file = 'stock_data.xlsx'
workbook = load_workbook(file)
worksheet = workbook.active
important_stuff = {}

for col in worksheet.iter_cols(values_only=True):
    stock_name = col[0]
    if stock_name is None:
        continue

    most_recent_value = None

    for row in col[1:]:
        if row is not None:
            try:
                items = row.split(";")
                for item in items:
                    open_price, current_price, change = item.split(", ")
                    most_recent_value = {
                        "Open": open_price.strip(),
                        "Current": current_price.strip(),
                        "Change": change.strip() + '%'
                    }
            except ValueError:
                pass

    if most_recent_value:
        important_stuff[stock_name] = most_recent_value

workbook.close()

def read_info(stock_name):
    if stock_name in list(important_stuff.keys()):
        return important_stuff[stock_name]
    else:
        return None

def read_all_info():
    for count in range(0, len(important_stuff.keys())):
        stock_name = list(important_stuff.keys())[count]
        print(f"Stock Name: {stock_name}")
        print(f"Open Price: {important_stuff[stock_name]['Open']}")
        print(f"Current Price: {important_stuff[stock_name]['Current']}")
        print(f"Change: {important_stuff[stock_name]['Change']}")
        print("-" * 20)
        time.sleep(1.5)

def main():
    while True:
        try:
            #choice = int(input("\nEnter 1 to read specific stock info, 2 to read all info, or 0 to quit: \n"))
            choice = 2
        except ValueError:
            print("Invalid input. Please enter a number.")
            continue

        if choice == 1:
            stock_name = input("Enter the stock name: \n").upper()
            info = read_info(stock_name)
            if info:
                print(f"Stock Name: {stock_name}")
                print(f"Open Price: {info['Open']}")
                print(f"Current Price: {info['Current']}")
                print(f"Change: {info['Change']}")
            else:
                print("Stock not found.")
            continue
        elif choice == 2:
            print(read_all_info())
            time.sleep(3600)
            continue
        elif choice == 0:
            break
        else:
            print("Invalid choice. Please enter 1, 2, or 0.")
if __name__ == "__main__":
    main()